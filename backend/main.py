from fastapi import FastAPI, Form, UploadFile, HTTPException, Query, Path
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional, Dict, Any, Literal
from pydantic import BaseModel, Field
from datetime import datetime
import uvicorn
import boto3
import os
import dotenv
from pymongo import AsyncMongoClient
from bson import ObjectId
from pydantic import BaseModel, Field
from pydantic_core import core_schema
from pymongo.errors import ServerSelectionTimeoutError
import json
from contextlib import contextmanager
import tempfile
import ifcopenshell
from collections import Counter
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import Response

dotenv.load_dotenv()

# AWS Configuration
AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_KEY")
S3_BUCKET = os.getenv("S3_BUCKET")
AWS_REGION = os.getenv("AWS_REGION", "ap-southeast-2")

# MongoDB Configuration
MONGODB_URL = os.getenv("MONGODB_URL", "mongodb://localhost:27017")
DB_NAME = os.getenv("DB_NAME", "project_management")

app = FastAPI(title="Project Management API")

app.mongodb_client = AsyncMongoClient(MONGODB_URL)
app.mongodb = app.mongodb_client[DB_NAME]

# CORS middleware configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow requests from your frontend
    allow_credentials=True,  # Allow cookies and authentication headers
    allow_methods=["*"],  # Allow all HTTP methods (GET, POST, etc.)
    allow_headers=["*"],  # Allow all headers (e.g., Authorization)
)


# S3 Client
s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY,
    region_name=AWS_REGION,
)


class PyObjectId(str):
    @classmethod
    def __get_pydantic_core_schema__(
        cls, _source_type: Any, _handler: Any
    ) -> core_schema.CoreSchema:
        return core_schema.json_or_python_schema(
            json_schema=core_schema.str_schema(),
            python_schema=core_schema.union_schema(
                [
                    core_schema.is_instance_schema(ObjectId),
                    core_schema.chain_schema(
                        [
                            core_schema.str_schema(),
                            core_schema.no_info_plain_validator_function(cls.validate),
                        ]
                    ),
                ]
            ),
            serialization=core_schema.plain_serializer_function_ser_schema(
                lambda x: str(x)
            ),
        )

    @classmethod
    def validate(cls, value) -> ObjectId:
        if not ObjectId.is_valid(value):
            raise ValueError("Invalid ObjectId")

        return ObjectId(value)


class UserPermissions(BaseModel):
    role: str
    permissions: List[str]


class IFCVersion(BaseModel):
    total_ec: float
    date_uploaded: datetime
    uploaded_by: str
    file_path: str


class EditHistory(BaseModel):
    timestamp: datetime
    user: str
    action: str
    description: str


class Project(BaseModel):
    id: PyObjectId = Field(default_factory=PyObjectId, alias="_id")
    project_name: str
    client_name: str
    typology: Optional[str] = None
    status: Optional[str] = None
    last_edited_date: datetime
    last_edited_user: str
    user_job_role: str
    current_version: int
    benchmark: Dict[str, int]
    access_control: Dict[str, UserPermissions]
    edit_history: List[EditHistory]
    ifc_versions: Dict[str, IFCVersion]

    class Config:
        populate_by_name = True
        arbitrary_types_allowed = True
        json_encoders = {ObjectId: str}


class Material(BaseModel):
    material: str
    ec: float


class Element(BaseModel):
    element: str
    ec: float
    materials: List[Material]


class Category(BaseModel):
    category: str
    total_ec: float
    elements: List[Element]


class ECBreakdown(BaseModel):
    total_ec: float
    ec_breakdown: List[Category]


class ProjectBreakdown(BaseModel):
    project_id: PyObjectId = Field(default_factory=PyObjectId, alias="_id")
    gfa: float
    summary: dict
    ec_breakdown: ECBreakdown
    last_calculated: datetime
    version: str


class ProjectBasicInfo(BaseModel):
    project_id: PyObjectId = Field(default_factory=PyObjectId, alias="_id")
    project_name: str
    client_name: str
    typology: str
    status: str
    benchmark: Dict[str, int]
    latest_version: str
    gfa: float
    file_path: str


class VersionHistory(BaseModel):
    version: str
    uploaded_by: str
    date_uploaded: datetime
    comments: str
    status: str
    total_ec: float
    gfa: float


class ProjectHistoryResponse(BaseModel):
    history: List


class Material(BaseModel):
    id: PyObjectId = Field(default_factory=PyObjectId, alias="_id")
    material_type: str
    specified_material: str
    density: Optional[float] = None
    embodied_carbon: float
    unit: Literal["kg", "m2"]
    database_source: Literal["Custom", "System"]
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.now)
    count: int = 0

    class Config:
        populate_by_name = True
        arbitrary_types_allowed = True
        json_encoders = {ObjectId: str}


class MaterialCreate(BaseModel):
    material_type: str
    specified_material: str
    density: Optional[float] = None
    embodied_carbon: float
    unit: Literal["kg", "m2"]
    database_source: Literal["Custom", "System"] = "Custom"


@contextmanager
def temp_ifc_file(content: bytes):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".ifc")
    try:
        tmp.write(content)
        tmp.close()
        yield tmp.name
    finally:
        os.unlink(tmp.name)


@app.post("/projects/{project_id}/materials", response_model=Material)
async def upload_material_and_queue(
    project_id: str,
    material: MaterialCreate,
    user_id: str = Query(..., description="ID of the user adding the material"),
    version: Optional[str] = Query(
        None,
        description="Project version to update. If not provided, uses current version",
    ),
):
    """
    Add a new material to the database and queue a recalculation of the project's EC values.

    Parameters:
    - project_id: The ID of the project to update
    - material: The material data to add
    - user_id: ID of the user adding the material
    - version: Optional version number. If not provided, uses current version

    Returns:
    - The newly created material
    """
    # Find the project
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(
            status_code=404, detail=f"Project with ID {project_id} not found"
        )

    # Determine which version to use
    version_number = version if version else str(project.get("current_version"))
    if version_number not in project["ifc_versions"]:
        raise HTTPException(
            status_code=404, detail=f"Version {version_number} not found"
        )

    # Check if a material with the same type and specification already exists
    existing_material = await app.mongodb.materials.find_one(
        {
            "material_type": material.material_type,
            "specified_material": material.specified_material,
        }
    )

    if existing_material:
        raise HTTPException(
            status_code=400,
            detail=f"Material '{material.specified_material}' of type '{material.material_type}' already exists",
        )

    # Create new material document
    new_material = Material(
        material_type=material.material_type,
        specified_material=material.specified_material,
        density=material.density,
        embodied_carbon=material.embodied_carbon,
        unit=material.unit,
        database_source=material.database_source,
        created_by=user_id,
        created_at=datetime.now(),
    )

    # Insert into database
    result = await app.mongodb.materials.insert_one(
        new_material.dict(by_alias=True, exclude={"id"})
    )

    # Retrieve created material
    created_material = await app.mongodb.materials.find_one({"_id": result.inserted_id})

    # Get the S3 path for the current version
    ifc_data = project["ifc_versions"].get(version_number, {})
    s3_path = ifc_data.get("file_path", "")

    if not s3_path:
        raise HTTPException(
            status_code=400, detail=f"No IFC file found for version {version_number}"
        )

    # Create a message for the SQS queue to trigger recalculation
    message = {
        "project_id": project_id,
        "ifc_version": version_number,
        "s3_path": s3_path,
        "user_id": user_id,
        "trigger": "material_added",
        "material_id": str(result.inserted_id),
        "timestamp": datetime.now().isoformat(),
    }

    # Send message to SQS queue
    try:
        sqs_client = boto3.client(
            "sqs",
            aws_access_key_id=AWS_ACCESS_KEY,
            aws_secret_access_key=AWS_SECRET_KEY,
            region_name=AWS_REGION,
        )
        queue_url = os.environ.get("SQS_QUEUE_URL")

        response = sqs_client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps(message),
            MessageGroupId=project_id,  # Group messages by project ID
            MessageDeduplicationId=f"{project_id}-{version_number}-material-{result.inserted_id}",  # Ensure idempotency
        )

        # Update project status to indicate recalculation is queued
        await app.mongodb.projects.update_one(
            {"_id": ObjectId(project_id)},
            {
                "$set": {
                    f"ifc_versions.{version_number}.calculation_status": "queued",
                    "last_edited_date": datetime.now(),
                    "last_edited_user": user_id,
                },
                "$push": {
                    "edit_history": {
                        "timestamp": datetime.now(),
                        "user": user_id,
                        "action": "material_added",
                        "description": f"Added material '{material.specified_material}' and queued recalculation for version {version_number}",
                    }
                },
            },
        )

    except Exception as e:
        # Even if queue fails, we've already added the material, so log the error but don't fail completely
        print(f"Error queuing recalculation: {str(e)}")
        # Add error note to project history
        await app.mongodb.projects.update_one(
            {"_id": ObjectId(project_id)},
            {
                "$push": {
                    "edit_history": {
                        "timestamp": datetime.now(),
                        "user": user_id,
                        "action": "error",
                        "description": f"Failed to queue recalculation after adding material: {str(e)}",
                    }
                },
            },
        )

    return created_material


@app.get("/projects/{project_id}/get_building_elements")
async def get_building_elements(
    project_id: str,
    version: Optional[str] = Query(
        None,
        description="IFC version to analyze. If not provided, uses current version",
    ),
):
    """
    Retrieve building elements with their material information and return as Excel file.
    """
    try:

        # Find the project
        project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

        if not project:
            raise HTTPException(
                status_code=404, detail=f"Project with ID {project_id} not found"
            )

        # Determine which version to use
        version_number = version if version else str(project.get("current_version"))
        if version_number not in project["ifc_versions"]:
            raise HTTPException(
                status_code=404, detail=f"Version {version_number} not found"
            )

        # Get the EC breakdown ID from the project
        ifc_version = project["ifc_versions"].get(version_number)
        ec_breakdown_id = ifc_version.get("ec_breakdown_id")
        calculation_status = ifc_version.get("calculation_status", "")

        if calculation_status != "completed":
            raise HTTPException(
                status_code=400,
                detail=f"EC calculation for version {version_number} is not completed. Current status: {calculation_status}",
            )

        if not ec_breakdown_id:
            raise HTTPException(
                status_code=404,
                detail=f"EC breakdown data not found for version {version_number}.",
            )

        # Get the EC breakdown data from MongoDB
        ec_breakdown = await app.mongodb.ec_breakdown.find_one({"_id": ec_breakdown_id})

        if not ec_breakdown:
            raise HTTPException(status_code=404, detail="EC breakdown data not found")

        # Extract the excel_data
        excel_data_rows = ec_breakdown.get("excel_data", [])

        if (
            not excel_data_rows or len(excel_data_rows) <= 1
        ):  # Check if there's data beyond headers
            raise HTTPException(
                status_code=404, detail="No building elements data found"
            )

        # Get headers and content
        headers = excel_data_rows[0]
        rows = excel_data_rows[1:]

        # Get materials data
        materials_data = await app.mongodb.materials.find().to_list(1000)

        # Create a workbook and sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Building Elements"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Map excel_data columns to our desired output format
        building_elements = []

        # Process each row and map to desired format
        for row in rows:
            if len(row) < 6:  # Make sure row has enough elements
                continue

            element_id = row[0]
            ifc_type = row[1]
            element_type = row[2]
            material = row[3]
            material_ec = row[4] if row[4] != "" else 0
            material_quantity = row[5] if row[5] != "" else 0
            units = row[6]

            # Find material information from materials_data
            material_info = None
            for m in materials_data:
                if m.get("specified_material") == material:
                    material_info = m
                    break

            building_material_family = (
                material_info.get("material_type") if material_info else "Rebar"
            )
            material = material

            building_elements.append(
                {
                    "element_id": element_id,
                    "ifc_type": ifc_type,
                    "element_type": element_type,
                    "building_material_family": building_material_family,
                    "material": material,
                    "material_ec": material_ec,
                    "material_quantity": material_quantity,
                    "units": units,
                }
            )

        # Add headers to building elements sheet
        element_headers = [
            "Element ID",
            "IFC Type",
            "Element Type",
            "Material Type",
            "Material",
            "Material EC (kgCO2e)",
            "Material Quantity",
            "Units",
        ]

        # Apply headers and formatting
        for col_num, header in enumerate(element_headers, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws1.column_dimensions[chr(64 + col_num)].width = 20  # Set column width

        # Add data rows for building elements
        for i, element in enumerate(building_elements, 2):
            ws1.cell(row=i, column=1, value=element["element_id"])
            ws1.cell(row=i, column=2, value=element["ifc_type"])
            ws1.cell(row=i, column=3, value=element["element_type"])
            ws1.cell(row=i, column=4, value=element["building_material_family"])
            ws1.cell(row=i, column=5, value=element["material"])
            ws1.cell(row=i, column=6, value=element["material_ec"])
            ws1.cell(row=i, column=7, value=element["material_quantity"])
            ws1.cell(row=i, column=8, value=element["units"])

        # Create a second sheet for detected materials
        ws2 = wb.create_sheet(title="Detected Materials")
        # Add headers with formatting for the materials sheet
        material_headers = [
            "Material Type",
            "Material",
            "Density (kg/m3)",
            "A1-A3 Embodied Carbon Emission / Unit",
            "Unit",
            "Data Source",
        ]
        # Apply headers and formatting
        for col_num, header in enumerate(material_headers, 1):
            cell = ws2.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            ws2.column_dimensions[chr(64 + col_num)].width = (
                25  # Set wider column width
            )

        # Get unique materials from the building elements
        unique_materials = {}
        for element in building_elements:
            if element["building_material_family"] and element["material"]:
                material_key = (
                    f"{element['building_material_family']}_{element['material']}"
                )
                if material_key not in unique_materials:
                    # Find this material in the materials database
                    material_data = None
                    for m in materials_data:
                        if (
                            m.get("material_type")
                            == element["building_material_family"]
                            and m.get("specified_material") == element["material"]
                        ):
                            material_data = m
                            break
                    if material_data:
                        unique_materials[material_key] = {
                            "family": element["building_material_family"],
                            "type": element["material"],
                            "density": (
                                material_data.get("density")
                                if material_data.get("density") != None
                                else "-"
                            ),
                            "unit": material_data.get("unit", "kg"),
                            "embodied_carbon": material_data.get("embodied_carbon"),
                            "source": material_data.get("database_source", "System"),
                        }

        # Add material data rows
        for i, (_, material) in enumerate(unique_materials.items(), 2):
            ws2.cell(row=i, column=1, value=material["family"])
            ws2.cell(row=i, column=2, value=material["type"])
            ws2.cell(row=i, column=3, value=material["density"])
            ws2.cell(row=i, column=4, value=material["embodied_carbon"])
            ws2.cell(row=i, column=5, value=material["unit"])
            ws2.cell(row=i, column=6, value=material["source"])

        for col_num, header in enumerate(element_headers, 1):
            col_letter = get_column_letter(col_num)
            # Set width based on header length plus padding
            width = len(header) + 5  # Add padding
            width = max(10, min(width, 50))  # Min 10, max 50
            ws1.column_dimensions[col_letter].width = width

        # For Materials sheet
        for col_num, header in enumerate(material_headers, 1):
            col_letter = get_column_letter(col_num)
            # Set width based on header length plus padding
            width = len(header) + 5  # Add padding
            width = max(10, min(width, 50))  # Min 10, max 50
            ws2.column_dimensions[col_letter].width = width
        # Save to a BytesIO object
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        # Return the Excel file as a response
        return Response(
            content=excel_output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=building_elements_{project_id}_v{version_number}.xlsx"
            },
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error processing building elements: {str(e)}"
        )


@app.get("/projects/{project_id}/{version_id}/calculation_status", response_model=Dict[str, str])
async def get_calculation_status(
    project_id: str,
    version_id: str,
    calculation_type: str = Query(None, description="Type of calculation to check: 'standard' or 'ai'. If not provided, returns both.")
):
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Find the specific version using version_id
    if "ifc_versions" not in project or str(version_id) not in project["ifc_versions"]:
        raise HTTPException(status_code=404, detail="Version not found")
    
    version = project["ifc_versions"][str(version_id)]
    
    # Prepare the response
    result = {}
    
    # Check for specific calculation type if requested
    if calculation_type == "standard":
        if "calculation_status" in version:
            result["standard"] = version["calculation_status"]
        else:
            result["standard"] = "unknown"
            
    elif calculation_type == "ai":
        if "ai_calculation_status" in version:
            result["ai"] = version["ai_calculation_status"]
        else:
            result["ai"] = "unknown"
            
    # If no specific type requested, return status for both
    else:
        # Standard calculation status
        if "calculation_status" in version:
            result["standard"] = version["calculation_status"]
        else:
            result["standard"] = "unknown"
            
        # AI calculation status
        if "ai_calculation_status" in version:
            result["ai"] = version["ai_calculation_status"]
        else:
            result["ai"] = "unknown"
    
    if not result:
        raise HTTPException(
            status_code=404, 
            detail="Calculation status not found for this version"
        )
        
    return result

@app.get("/projects/{project_id}/missing_materials", response_model=Dict[str, Any])
async def get_missing_materials(
    project_id: str,
    version: Optional[str] = Query(
        None,
        description="IFC version to analyze. If not provided, uses current version",
    ),
):
    """
    Retrieve missing materials detected in the IFC model.
    Returns the type of elements, element ID, and error type for materials not found in the system database.
    """
    # Find the project
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(
            status_code=404, detail=f"Project with ID {project_id} not found"
        )

    # Determine which version to use
    version_number = version if version else str(project.get("current_version"))
    if version_number not in project["ifc_versions"]:
        raise HTTPException(
            status_code=404, detail=f"Version {version_number} not found"
        )

    # Get the EC breakdown ID from the project
    ifc_version = project["ifc_versions"].get(version_number)
    ec_breakdown_id = ifc_version.get("ec_breakdown_id")
    calculation_status = ifc_version.get("calculation_status", "")

    if calculation_status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"EC calculation for version {version_number} is not completed. Current status: {calculation_status}",
        )

    if not ec_breakdown_id:
        raise HTTPException(
            status_code=404,
            detail=f"EC breakdown data not found for version {version_number}.",
        )

    # Get the EC breakdown data from MongoDB
    ec_breakdown = await app.mongodb.ec_breakdown.find_one({"_id": ec_breakdown_id})
    print("Get Ec breakdown")
    if not ec_breakdown:
        raise HTTPException(status_code=404, detail="EC breakdown data not found")

    # Extract missing materials from the EC breakdown data
    breakdown = ec_breakdown.get("breakdown")
    missing_materials = breakdown.get("missing_materials", {})

    # Format the response
    # Use a set to store unique specified materials
    unique_materials = set()
    result = []
    for ifc_type, materials in missing_materials.items():
        print("Get missing materials: ", ifc_type, materials)
        for id, material_data in materials:
            # Check if the material is already in the set
            if material_data == "Undefined":
                result.append(
                    {
                        "IfcType": ifc_type,
                        "ElementId": id,
                        "SpecifiedMaterial": material_data,
                        "ErrorType": "Material Undefined",
                    }
                )
                continue
            if material_data not in unique_materials:
                unique_materials.add(material_data)
                result.append(
                    {
                        "IfcType": ifc_type,
                        "ElementId": id,
                        "SpecifiedMaterial": material_data,
                        "ErrorType": "Material not found in system database",
                    }
                )

    return {
        "project_id": project_id,
        "version": version_number,
        "total_missing_materials": len(result),
        "missing_materials": result,
    }


@app.get("/projects/{project_id}/missing_elements", response_model=Dict[str, Any])
async def get_missing_elements(
    project_id: str,
    version: Optional[str] = Query(
        None,
        description="IFC version to analyze. If not provided, uses current version",
    ),
):
    """
    Retrieve missing elements detected in the IFC model.
    Returns the type of elements, element ID, and error type for elements not properly classified.
    """
    # Find the project
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(
            status_code=404, detail=f"Project with ID {project_id} not found"
        )

    # Determine which version to use
    version_number = version if version else str(project.get("current_version"))
    if version_number not in project["ifc_versions"]:
        raise HTTPException(
            status_code=404, detail=f"Version {version_number} not found"
        )

    # Get the EC breakdown ID from the project
    ifc_version = project["ifc_versions"].get(version_number)
    ec_breakdown_id = ifc_version.get("ec_breakdown_id")
    calculation_status = ifc_version.get("calculation_status", "")

    if calculation_status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"EC calculation for version {version_number} is not completed. Current status: {calculation_status}",
        )

    if not ec_breakdown_id:
        raise HTTPException(
            status_code=404,
            detail=f"EC breakdown data not found for version {version_number}.",
        )

    # Get the EC breakdown data from MongoDB
    ec_breakdown = await app.mongodb.ec_breakdown.find_one({"_id": ec_breakdown_id})

    if not ec_breakdown:
        raise HTTPException(status_code=404, detail="EC breakdown data not found")

    # Extract missing elements (element_type_skipped) from the EC breakdown data
    breakdown = ec_breakdown.get("breakdown")
    element_type_skipped = breakdown.get("element_type_skipped", [])

    # Format the response
    result = []
    for element_info in element_type_skipped:
        if isinstance(element_info, list) and len(element_info) >= 2:
            element_id, element_type = element_info
            result.append(
                {
                    "IfcType": element_type,
                    "ElementId": element_id,
                    "ErrorType": "MaterialElement not classified",
                }
            )

    return {
        "project_id": project_id,
        "version": version_number,
        "total_missing_elements": len(result),
        "missing_elements": result,
    }


@app.get("/ifc/elements", response_model=Dict[str, Any])
async def get_ifc_elements(
    ifc_path: str = Query(..., description="S3 path to the IFC file")
):
    """
    Retrieve elements detected in the IFC model.
    Returns the type of elements and their quantities.

    The frontend should provide the complete S3 path to the IFC file.
    """
    if not ifc_path or not ifc_path.startswith("s3://"):
        raise HTTPException(status_code=400, detail="Invalid IFC file path")

    # Extract the key from the S3 path
    s3_key = ifc_path.replace(f"s3://{S3_BUCKET}/", "")

    try:
        # Get the IFC file from S3
        response = s3_client.get_object(
            Bucket=S3_BUCKET,
            Key=s3_key,
        )
        file_content = response["Body"].read()

        element_counts = {}

        with temp_ifc_file(file_content) as ifc_file_path:
            ifc_file = ifcopenshell.open(ifc_file_path)

            # Get only the specified element types
            columns = ifc_file.by_type("IfcColumn")
            beams = ifc_file.by_type("IfcBeam")
            slabs = ifc_file.by_type("IfcSlab")
            walls = ifc_file.by_type("IfcWall")
            windows = ifc_file.by_type("IfcWindow")
            roofs = ifc_file.by_type("IfcRoof")
            doors = ifc_file.by_type("IfcDoor")
            stairs = ifc_file.by_type("IfcStairFlight")
            railings = ifc_file.by_type("IfcRailing")
            members = ifc_file.by_type("IfcMember")
            plates = ifc_file.by_type("IfcPlate")
            piles = ifc_file.by_type("IfcPile")
            footings = ifc_file.by_type("IfcFooting")
            spaces = ifc_file.by_type("IfcSpace")

            # Count each element type
            element_counts = {
                "IfcColumn": len(columns),
                "IfcBeam": len(beams),
                "IfcSlab": len(slabs),
                "IfcWall": len(walls),
                "IfcWindow": len(windows),
                "IfcRoof": len(roofs),
                "IfcDoor": len(doors),
                "IfcStairFlight": len(stairs),
                "IfcRailing": len(railings),
                "IfcMember": len(members),
                "IfcPlate": len(plates),
                "IfcPile": len(piles),
                "IfcFooting": len(footings),
                "IfcSpace": len(spaces),
            }

            # Remove any element types with zero count
            element_counts = {k: v for k, v in element_counts.items() if v > 0}

        sorted_elements = {
            k: v
            for k, v in sorted(
                element_counts.items(), key=lambda item: item[1], reverse=True
            )
        }

        return {
            "ifc_path": ifc_path,
            "elements": sorted_elements,
        }

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error processing IFC file: {str(e)}"
        )


@app.get("/materials", response_model=List[Material])
async def get_materials(
    project_id: Optional[str] = Query(
        None,
        description="Project ID to filter materials by. If not provided, returns all materials.",
    ),
    version: Optional[str] = Query(
        None,
        description="IFC version to analyze. If not provided, uses current version",
    ),
):
    """
    Retrieve a list of materials from MongoDB.
    If project_id is not provided, returns all materials from the database.
    If project_id is provided, returns materials for that specific project.
    If version is provided, it will return materials for that specific version.
    If version is not provided, it will use the current version of the project.
    The response includes the count of each material's occurrences in the project.
    """
    try:
        print(project_id)
        # If no project_id is provided, return all materials
        if not project_id:
            materials = await app.mongodb.materials.find().to_list(1000)
            # Convert ObjectId to string for each material
            for material in materials:
                material["_id"] = str(material["_id"])
                material["count"] = 0  # No project context, so count is 0
            return materials
        # Find the project
        project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})
        if not project:
            raise HTTPException(
                status_code=404, detail=f"Project with ID {project_id} not found"
            )
        # Determine which version to use
        version_number = version if version else str(project.get("current_version"))
        if version_number not in project["ifc_versions"]:
            raise HTTPException(
                status_code=404, detail=f"Version {version_number} not found"
            )
        # Get the EC breakdown ID from the project
        ifc_version = project["ifc_versions"].get(version_number)
        ec_breakdown_id = ifc_version.get("ec_breakdown_id")
        if not ec_breakdown_id:
            return []
        # Get the breakdown document
        breakdown = await app.mongodb.ec_breakdown.find_one(
            {"_id": ObjectId(ec_breakdown_id)}
        )
        if not breakdown:
            return []  # Breakdown not found
        
        if "material_counts" in breakdown and breakdown["material_counts"]:
            material_counts = breakdown["material_counts"]
            
            # Get all materials that are in the count dictionary
            materials = await app.mongodb.materials.find(
                {"specified_material": {"$in": list(material_counts.keys())}}
            ).to_list(1000)
            
            # Add counts to each material
            for material in materials:
                material["_id"] = str(material["_id"])
                material_name = material["specified_material"]
                material["count"] = material_counts.get(material_name, 0)
                
            return materials
        # Use Counter to track material frequencies
        material_counter = Counter()

        if "breakdown" in breakdown and "ec_breakdown" in breakdown["breakdown"]:
            for category in breakdown["breakdown"]["ec_breakdown"]:
                if "elements" in category:
                    for element in category["elements"]:
                        if "materials" in element:
                            for material in element["materials"]:
                                if "material" in material:
                                    material_counter[material["material"]] += 1

        if not material_counter:
            return []  # No material IDs found in breakdown

        print(dict(material_counter))

        # Query for materials using the collected IDs
        materials = await app.mongodb.materials.find(
            {"specified_material": {"$in": list(material_counter.keys())}}
        ).to_list(1000)

        # Convert ObjectId to string for each material and add count
        for material in materials:
            material["_id"] = str(material["_id"])
            material["count"] = material_counter[material["specified_material"]]

        return materials

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error retrieving materials: {str(e)}"
        )


@app.post("/materials", response_model=Material)
async def upload_material(
    material: MaterialCreate,
    user_id: str = Query(..., description="ID of the user adding the material"),
):
    """
    Add a new material to the database.
    """
    # Check if a material with the same type and specification already exists
    existing_material = await app.mongodb.materials.find_one(
        {
            "material_type": material.material_type,
            "specified_material": material.specified_material,
        }
    )

    if existing_material:
        raise HTTPException(
            status_code=400,
            detail=f"Material '{material.specified_material}' of type '{material.material_type}' already exists",
        )

    # Create new material document
    new_material = Material(
        material_type=material.material_type,
        specified_material=material.specified_material,
        density=material.density,
        embodied_carbon=material.embodied_carbon,
        unit=material.unit,
        database_source=material.database_source,
        created_by=user_id,
        created_at=datetime.now(),
    )

    # Insert into database
    result = await app.mongodb.materials.insert_one(
        new_material.dict(by_alias=True, exclude={"id"})
    )

    # Retrieve created material
    created_material = await app.mongodb.materials.find_one({"_id": result.inserted_id})

    return created_material


@app.delete("/materials/{material_id}", response_model=Dict[str, Any])
async def delete_material(
    material_id: str = Path(..., description="ID of the material to delete"),
    user_id: str = Query(..., description="ID of the user deleting the material"),
):
    """
    Delete a material from the database.
    """
    # Check if material exists
    material = await app.mongodb.materials.find_one({"_id": ObjectId(material_id)})

    if not material:
        raise HTTPException(
            status_code=404, detail=f"Material with ID {material_id} not found"
        )

    # Check if the material is a system default or created by the user
    # Optional: You might want to restrict deletion of system materials
    if (
        material.get("database_source") == "System"
        and material.get("created_by") != user_id
    ):
        raise HTTPException(
            status_code=403, detail="Cannot delete system-provided materials"
        )

    # Delete the material
    result = await app.mongodb.materials.delete_one({"_id": ObjectId(material_id)})

    if result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Failed to delete material")

    return {
        "success": True,
        "message": f"Material '{material.get('specified_material')}' deleted successfully",
        "material_id": material_id,
    }


@app.get("/test_db_connection")
async def test_db_connection():

    try:
        # List collections in the database to verify connectivity
        collections = await app.mongodb.list_collection_names()
        return {"success": True, "collections": collections}
    except ServerSelectionTimeoutError as e:
        return {"success": False, "error": str(e)}


@app.get("/projects", response_model=List[Project])
async def get_projects(
    user_id: str = Query(..., description="User ID to fetch projects for")
):
    """Get all projects accessible by a user"""
    print("User id ", user_id)
    projects = await app.mongodb.projects.find(
        {f"access_control.{user_id}": {"$exists": True}}
    ).to_list(1000)
    # print("Projects: ", projects)
    if not projects:
        raise HTTPException(
            status_code=404, detail=f"No projects found for user ID: {user_id}"
        )

    for project in projects:
        project["_id"] = str(project["_id"])

    return projects


@app.post("/projects/{project_id}/upload_ifc", response_model=Dict)
async def upload_ifc(
    project_id: str,
    file: UploadFile,
    user_id: str = Query(..., description="ID of the user uploading the file"),
    comments: str = Form(""),  # Default to empty string if not provided
    status: str = Form(""),
):
    if not file.filename.lower().endswith(".ifc"):
        raise HTTPException(
            status_code=400, detail="Invalid file format. Only IFC files are allowed."
        )

    # Verify user has upload permission
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})
    if not project:
        raise HTTPException(
            status_code=404, detail=f"Project with ID {project_id} not found"
        )

    user_permissions = project["access_control"].get(user_id, {}).get("permissions", [])
    if "upload" not in user_permissions:
        raise HTTPException(
            status_code=403, detail="User does not have permission to upload files"
        )

    try:
        file_content = await file.read()

        current_version = project.get("current_version")
        if current_version is None:
            new_version = "1"  # Start new projects with version "1"
        # elif current_version == 1:
        #     new_version = "1"  # Keep it as "1"
        else:
            new_version = str(current_version + 1)  # Increment for later versions

        s3_path = f"ifc_files/{project_id}/{new_version}_{file.filename}"

        # Upload to S3
        s3_client.put_object(
            Bucket=S3_BUCKET,
            Key=s3_path,
            Body=file_content,
            ContentType="application/octet-stream",
        )
        print("uploaded the file to s3")
        # Create a message for the SQS queue
        base_message = {
            "project_id": project_id,
            "ifc_version": new_version,
            "s3_path": f"s3://{S3_BUCKET}/{s3_path}",
            "user_id": user_id,
            "timestamp": datetime.now().isoformat(),
        }

        standard_message = base_message.copy()
        standard_message["calculation_type"] = "standard"
        standard_message["enable_ai_material_matcher"] = False

        # Send message to SQS queue
        sqs_client = boto3.client(
            "sqs",
            aws_access_key_id=AWS_ACCESS_KEY,
            aws_secret_access_key=AWS_SECRET_KEY,
            region_name="ap-southeast-2",
        )
        queue_url = os.environ.get("SQS_QUEUE_URL")

        sqs_client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps(standard_message),
            MessageGroupId=project_id,  # Group messages by project ID
            MessageDeduplicationId=f"{project_id}-{new_version}-standard",  # Ensure idempotency
        )
        print("uploaded standard calculation to queue")

        ai_message = base_message.copy()
        ai_message["enable_ai_material_matcher"] = True
        ai_message["calculation_type"] = "ai_enhanced"

        sqs_client.send_message(
            QueueUrl=queue_url,
            MessageBody=json.dumps(ai_message),
            MessageGroupId=project_id,  # Group messages by project ID
            MessageDeduplicationId=f"{project_id}-{new_version}-ai_enhanced",  # Ensure idempotency
        )
        print("uploaded both messages to queue")

        # Update MongoDB
        update_result = await app.mongodb.projects.update_one(
            {"_id": ObjectId(project_id)},
            {
                "$set": {
                    "current_version": int(new_version),
                    "last_edited_date": datetime.now(),
                    "last_edited_user": user_id,
                    f"ifc_versions.{new_version}": {
                        "date_uploaded": datetime.now(),
                        "uploaded_by": user_id,
                        "comments": comments,
                        "status": status,
                        "file_path": f"s3://{S3_BUCKET}/{s3_path}",
                        "gfa": 0,
                        "total_ec": 0,
                        "ai_total_ec" : 0
                        "calculation_status": "queued",
                        "ai_calculation_status": "queued",
                    },
                },
                "$push": {
                    "edit_history": {
                        "timestamp": datetime.now(),
                        "user": user_id,
                        "action": "uploaded_ifc",
                        "description": f"Uploaded version {new_version}",
                    }
                },
            },
        )

        return {
            "success": True,
            "message": "File uploaded successfully",
            "version": new_version,
            "s3_path": f"s3://{S3_BUCKET}/{s3_path}",
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload file: {str(e)}")


# Get EC breakdown and ec value
@app.get("/projects/{project_id}/get_breakdown", response_model=ProjectBreakdown)
async def get_breakdown(
    project_id: str, 
    version: str = None,
    calculation_type: Optional[str] = Query("standard", description="Type of calculation to retrieve: 'standard' or 'ai_enhanced'")
    ):
 
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    version_number = version if version else str(project.get("current_version"))
    if version_number not in project["ifc_versions"]:
        raise HTTPException(status_code=400, detail="Version not found")
    ifc_data = project["ifc_versions"].get(version_number, {})

    # Retrieve stored EC values and breakdowns
    total_ec = ifc_data.get("total_ec", 0)
    gfa = ifc_data.get("gfa", 0)

    ec_breakdown_id = ifc_data.get("ec_breakdown_id") if calculation_type == "standard" else ifc_data.get("ai_ec_breakdown_id")
    calculation_status = ifc_data.get("calculation_status") if calculation_type == "standard" else ifc_data.get("ai_calculation_status")
    print(ec_breakdown_id)
    
    if not ec_breakdown_id:
        raise HTTPException(
            status_code=404, 
            detail=f"EC breakdown for {calculation_type} calculation not found for version {version_number}"
        )
    
    if calculation_status != "completed":
        raise HTTPException(
            status_code=400,
            detail=f"{calculation_type.capitalize()} calculation for version {version_number} is not completed. Current status: {calculation_status}"
        )
    ec_breakdown_data = await app.mongodb.ec_breakdown.find_one(
        {"_id": ec_breakdown_id}
    )

    print("ec breakdown summary is,", ec_breakdown_data["summary"])
    print("ec_breakdown_data[breakdown] is,", ec_breakdown_data["breakdown"])
    return ProjectBreakdown(
        project_id=str(project["_id"]),
        gfa=gfa,
        summary=ec_breakdown_data["summary"],
        ec_breakdown=ec_breakdown_data["breakdown"],
        # maybe add the matched material..
        last_calculated=project.get("last_calculated", datetime.now()),
        version=version_number,
    )


@app.get("/projects/{project_id}/get_project_info", response_model=ProjectBasicInfo)
async def get_project_info(project_id: str):
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get the latest version's GFA
    latest_version = str(project.get("current_version"))
    ifc_data = project["ifc_versions"].get(latest_version, {})
    gfa = ifc_data.get("gfa", 0)
    file_path = ifc_data.get("file_path", "")
    print("Project data:", project)
    print("Benchmark data:", project.get("benchmark", "No benchmark field found."))

    # Extract benchmark values directly from the flat dictionary
    benchmark_values = project.get("benchmark", {})
    print("Flat benchmark data extracted:", benchmark_values)

    return ProjectBasicInfo(
        _id=project["_id"],
        project_name=project.get("project_name"),
        client_name=project.get("client_name"),
        typology=project.get("typology", "Not Specified"),
        status=project.get("status", "Not Specified"),
        benchmark=benchmark_values,
        latest_version=latest_version,
        gfa=gfa,
        file_path=file_path,
    )


# Get latest edits history (top 4 history, get the uploaded_by, date_uploaded, comments, status)
@app.get("/projects/{project_id}/get_history", response_model=ProjectHistoryResponse)
async def get_project_history(project_id: str):
    project = await app.mongodb.projects.find_one({"_id": ObjectId(project_id)})

    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get the last 4 versions uploaded
    ifc_versions = project.get("ifc_versions", {})
    sorted_versions = sorted(ifc_versions.keys(), key=lambda v: int(v), reverse=True)

    version_history = []
    for version in sorted_versions:
        version_data = ifc_versions.get(version, {})
        version_history.append(
            VersionHistory(
                version=version,
                uploaded_by=version_data.get("uploaded_by", ""),
                date_uploaded=version_data.get("date_uploaded", datetime.now()),
                comments=version_data.get("comments", ""),
                status=version_data.get("status", ""),
                total_ec=version_data.get("total_ec", 0.0),
                gfa=version_data.get("gfa", 0.0),
            )
        )

    return ProjectHistoryResponse(history=version_history)


@app.post("/create_project", response_model=Project)
async def create_project(project: Project):
    new_project = await app.mongodb.projects.insert_one(
        project.dict(by_alias=True, exclude={"id"})
    )
    created_project = await app.mongodb.projects.find_one(
        {"_id": new_project.inserted_id}
    )
    return created_project


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
